#!/usr/bin/env python3
"""
Análisis Completo de Clientes - Trade Unity

Genera Excel detallado con:
- TOP 100 clientes (facturación, reincidencia, salud)
- Segmentación RFV completa
- Análisis de comportamiento
- Métricas de retención y crecimiento
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
import os

try:
    import pandas as pd
    from openpyxl import load_workbook
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# Archivos
VENTAS_CSV = "inputs/ventas_historicas_items_FINAL.csv"
OUTPUT_DIR = "outputs"
OUTPUT_EXCEL = f"{OUTPUT_DIR}/TradeUnity Customer Intelligence.xlsx"


def parse_decimal(value: str) -> float:
    """Convierte string a float."""
    if not value or value == "":
        return 0.0
    
    value_str = str(value).strip().replace("$", "").replace(" ", "").replace("%", "")
    value_str = value_str.replace(",", ".")
    
    try:
        return float(value_str)
    except (ValueError, TypeError):
        return 0.0


def auto_adjust_column_widths(writer, sheet_name, df):
    """Ajusta automáticamente el ancho de las columnas en Excel."""
    try:
        worksheet = writer.sheets[sheet_name]
        
        for idx, col in enumerate(df.columns, 1):
            column_letter = worksheet.cell(row=1, column=idx).column_letter
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        print(f"   ⚠️  No se pudo ajustar columnas en {sheet_name}: {e}")


def load_ventas():
    """Carga datos de ventas."""
    print("📖 Cargando datos de ventas...")
    
    if not os.path.exists(VENTAS_CSV):
        print(f"   ⚠️  Archivo de ventas no encontrado: {VENTAS_CSV}")
        return pd.DataFrame()
    
    df = pd.read_csv(VENTAS_CSV, encoding='utf-8-sig')
    
    # Convertir fechas
    if 'Fecha Creación' in df.columns:
        df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación'], errors='coerce')
    
    # Convertir numéricos
    numeric_cols = ['Total Item con IVA', 'Total Orden', 'Cantidad Unitarias']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '.').str.replace('$', '').str.strip()
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    print(f"   ✅ {len(df)} registros de ventas cargados")
    return df


def analyze_clients(ventas_df):
    """Analiza clientes y genera métricas completas desde perspectiva CMO."""
    print("📊 Analizando clientes (perspectiva CMO)...")
    
    if len(ventas_df) == 0:
        return pd.DataFrame()
    
    # Preparar datos numéricos
    ventas_df['Descuento % Item'] = pd.to_numeric(
        ventas_df['Descuento % Item'].astype(str).str.replace('%', '').str.replace(',', '.'), 
        errors='coerce'
    ).fillna(0)
    ventas_df['Precio Original'] = pd.to_numeric(
        ventas_df['Precio Original'].astype(str).str.replace(',', '.').str.replace('$', ''), 
        errors='coerce'
    ).fillna(0)
    ventas_df['Precio Venta'] = pd.to_numeric(
        ventas_df['Precio Venta'].astype(str).str.replace(',', '.').str.replace('$', ''), 
        errors='coerce'
    ).fillna(0)
    ventas_df['Días desde Última Recepción CEG'] = pd.to_numeric(
        ventas_df['Días desde Última Recepción CEG'].astype(str).str.replace(',', '.'), 
        errors='coerce'
    ).fillna(0)
    
    # Preparar datos de márgenes y costos
    ventas_df['% Margen sobre FOB'] = pd.to_numeric(
        ventas_df['% Margen sobre FOB'].astype(str).str.replace('%', '').str.replace(',', '.'), 
        errors='coerce'
    ).fillna(0)
    ventas_df['% Margen sobre Plataforma'] = pd.to_numeric(
        ventas_df['% Margen sobre Plataforma'].astype(str).str.replace('%', '').str.replace(',', '.'), 
        errors='coerce'
    ).fillna(0)
    ventas_df['FOB CEG'] = pd.to_numeric(
        ventas_df['FOB CEG'].astype(str).str.replace(',', '.').str.replace('$', ''), 
        errors='coerce'
    ).fillna(0)
    ventas_df['Base Price CEG'] = pd.to_numeric(
        ventas_df['Base Price CEG'].astype(str).str.replace(',', '.').str.replace('$', ''), 
        errors='coerce'
    ).fillna(0)
    
    # Calcular porcentaje de compra sobre FOB y Plataforma
    ventas_df['Precio_Venta_Unitario'] = pd.to_numeric(
        ventas_df['Precio Venta Unitario'].astype(str).str.replace(',', '.').str.replace('$', ''), 
        errors='coerce'
    ).fillna(0)
    
    # % sobre FOB = (Precio Venta Unitario / FOB) - 1 * 100
    mask_fob = ventas_df['FOB CEG'] > 0
    ventas_df.loc[mask_fob, '%_Compra_Sobre_FOB'] = (
        (ventas_df.loc[mask_fob, 'Precio_Venta_Unitario'] / ventas_df.loc[mask_fob, 'FOB CEG'] - 1) * 100
    )
    ventas_df['%_Compra_Sobre_FOB'] = ventas_df['%_Compra_Sobre_FOB'].fillna(0)
    
    # % sobre Plataforma = (Precio Venta Unitario / Base Price) - 1 * 100
    mask_plat = ventas_df['Base Price CEG'] > 0
    ventas_df.loc[mask_plat, '%_Compra_Sobre_Plataforma'] = (
        (ventas_df.loc[mask_plat, 'Precio_Venta_Unitario'] / ventas_df.loc[mask_plat, 'Base Price CEG'] - 1) * 100
    )
    ventas_df['%_Compra_Sobre_Plataforma'] = ventas_df['%_Compra_Sobre_Plataforma'].fillna(0)
    
    # Calcular descuento si no existe
    ventas_df['Descuento_Calculado'] = 0.0
    mask = (ventas_df['Precio Original'] > 0) & (ventas_df['Precio Venta'] < ventas_df['Precio Original'])
    ventas_df.loc[mask, 'Descuento_Calculado'] = (
        (ventas_df.loc[mask, 'Precio Original'] - ventas_df.loc[mask, 'Precio Venta']) / 
        ventas_df.loc[mask, 'Precio Original'] * 100
    ).astype(float)
    ventas_df['Descuento_Final'] = ventas_df[['Descuento % Item', 'Descuento_Calculado']].max(axis=1).astype(float)
    
    # Agrupar por cliente - métricas básicas
    clientes_stats = ventas_df.groupby('Email Cliente').agg({
        'Número de Orden': 'nunique',
        'Total Item con IVA': 'sum',
        'Cantidad Unitarias': 'sum',
        'SKU': 'nunique',
        'Fecha Creación': ['min', 'max'],
        'Categoría (2° Nivel)': lambda x: x.mode()[0] if len(x.mode()) > 0 else 'N/A',
        'Brand Name CEG': lambda x: x.mode()[0] if len(x.mode()) > 0 else 'N/A',
        'Nombre Cliente': 'first',
        'Apellido Cliente': 'first',
        'CUIT Cliente': 'first'
    }).reset_index()
    
    clientes_stats.columns = ['Email', 'Ordenes', 'LTV', 'Unidades_Totales', 'SKUs_Unicos', 
                              'Primera_Compra', 'Ultima_Compra', 'Categoria_Favorita', 
                              'Marca_Favorita', 'Nombre', 'Apellido', 'CUIT']
    
    # Calcular métricas adicionales básicas
    clientes_stats['Dias_Activo'] = (clientes_stats['Ultima_Compra'] - clientes_stats['Primera_Compra']).dt.days
    clientes_stats['Dias_Desde_Ultima'] = (datetime.now() - clientes_stats['Ultima_Compra']).dt.days
    clientes_stats['Ticket_Promedio'] = clientes_stats['LTV'] / clientes_stats['Ordenes'].clip(lower=1)
    clientes_stats['Frecuencia_Mensual'] = clientes_stats['Ordenes'] / (clientes_stats['Dias_Activo'] / 30).clip(lower=1)
    clientes_stats['Reincidente'] = clientes_stats['Ordenes'] > 1
    clientes_stats['Cliente_Sano'] = (clientes_stats['Dias_Desde_Ultima'] <= 90) & (clientes_stats['Ordenes'] >= 2)
    
    # ========== ANÁLISIS DE MARKETING (CMO) ==========
    print("   🎯 Calculando métricas de marketing...")
    
    # 1. Fidelidad a Marcas (Fan de Marca)
    marca_stats = ventas_df.groupby(['Email Cliente', 'Brand Name CEG']).agg({
        'Total Item con IVA': 'sum',
        'Cantidad Unitarias': 'sum'
    }).reset_index()
    marca_stats = marca_stats.sort_values(['Email Cliente', 'Total Item con IVA'], ascending=[True, False])
    
    marca_dominante = marca_stats.groupby('Email Cliente').first().reset_index()
    marca_dominante.columns = ['Email', 'Marca_Dominante', 'Facturacion_Marca_Dominante', 'Unidades_Marca_Dominante']
    
    total_por_cliente = ventas_df.groupby('Email Cliente').agg({
        'Total Item con IVA': 'sum',
        'Brand Name CEG': 'nunique'
    }).reset_index()
    total_por_cliente.columns = ['Email', 'Facturacion_Total', 'Marcas_Unicas']
    
    fidelidad_marca = marca_dominante.merge(total_por_cliente, on='Email')
    fidelidad_marca['%_Facturacion_Marca_Dominante'] = (
        fidelidad_marca['Facturacion_Marca_Dominante'] / fidelidad_marca['Facturacion_Total'] * 100
    )
    fidelidad_marca['Es_Fan_Marca'] = fidelidad_marca['%_Facturacion_Marca_Dominante'] >= 70
    fidelidad_marca['Fan_Marca_Nombre'] = fidelidad_marca.apply(
        lambda x: x['Marca_Dominante'] if x['Es_Fan_Marca'] else 'Diversificado', axis=1
    )
    
    clientes_stats = clientes_stats.merge(
        fidelidad_marca[['Email', 'Marca_Dominante', '%_Facturacion_Marca_Dominante', 
                         'Es_Fan_Marca', 'Fan_Marca_Nombre', 'Marcas_Unicas']], 
        on='Email', how='left'
    )
    
    # 2. Fidelidad a Vertical/Categoría
    categoria_stats = ventas_df.groupby(['Email Cliente', 'Categoría (2° Nivel)']).agg({
        'Total Item con IVA': 'sum',
        'Cantidad Unitarias': 'sum'
    }).reset_index()
    categoria_stats = categoria_stats.sort_values(['Email Cliente', 'Total Item con IVA'], ascending=[True, False])
    
    categoria_dominante = categoria_stats.groupby('Email Cliente').first().reset_index()
    categoria_dominante.columns = ['Email', 'Categoria_Dominante', 'Facturacion_Categoria_Dominante', 'Unidades_Categoria_Dominante']
    
    total_categorias = ventas_df.groupby('Email Cliente').agg({
        'Categoría (2° Nivel)': 'nunique'
    }).reset_index()
    total_categorias.columns = ['Email', 'Categorias_Unicas']
    
    fidelidad_categoria = categoria_dominante.merge(total_por_cliente[['Email', 'Facturacion_Total']], on='Email')
    fidelidad_categoria = fidelidad_categoria.merge(total_categorias, on='Email')
    fidelidad_categoria['%_Facturacion_Categoria_Dominante'] = (
        fidelidad_categoria['Facturacion_Categoria_Dominante'] / fidelidad_categoria['Facturacion_Total'] * 100
    )
    fidelidad_categoria['Es_Fiel_Vertical'] = fidelidad_categoria['%_Facturacion_Categoria_Dominante'] >= 60
    fidelidad_categoria['Fiel_Vertical_Nombre'] = fidelidad_categoria.apply(
        lambda x: x['Categoria_Dominante'] if x['Es_Fiel_Vertical'] else 'Diversificado', axis=1
    )
    
    clientes_stats = clientes_stats.merge(
        fidelidad_categoria[['Email', 'Categoria_Dominante', '%_Facturacion_Categoria_Dominante',
                              'Es_Fiel_Vertical', 'Fiel_Vertical_Nombre', 'Categorias_Unicas']],
        on='Email', how='left'
    )
    
    # 3. Diversidad de Compra (Salpicado vs Especializado)
    clientes_stats['Diversidad_Compra'] = clientes_stats.apply(
        lambda x: 'Muy Diversificado' if (x.get('Categorias_Unicas', 0) >= 5 and x.get('Marcas_Unicas', 0) >= 5)
        else 'Diversificado' if (x.get('Categorias_Unicas', 0) >= 3 or x.get('Marcas_Unicas', 0) >= 3)
        else 'Especializado', axis=1
    )
    
    # 4. Análisis de Descuentos
    descuento_stats = ventas_df.groupby('Email Cliente').agg({
        'Descuento_Final': ['mean', 'max', lambda x: (x > 0).sum()],
        'Total Item con IVA': 'sum',
        'Cantidad Unitarias': 'sum'
    }).reset_index()
    descuento_stats.columns = ['Email', 'Descuento_Promedio_%', 'Descuento_Maximo_%', 'Items_Con_Descuento', 
                                'Facturacion_Total_Descuento', 'Unidades_Con_Descuento']
    
    total_items = ventas_df.groupby('Email Cliente').size().reset_index()
    total_items.columns = ['Email', 'Total_Items']
    descuento_stats = descuento_stats.merge(total_items, on='Email')
    descuento_stats['%_Items_Con_Descuento'] = (
        descuento_stats['Items_Con_Descuento'] / descuento_stats['Total_Items'] * 100
    )
    descuento_stats['Es_Cazador_Descuentos'] = (
        (descuento_stats['Descuento_Promedio_%'] >= 15) | 
        (descuento_stats['%_Items_Con_Descuento'] >= 50)
    )
    
    clientes_stats = clientes_stats.merge(
        descuento_stats[['Email', 'Descuento_Promedio_%', 'Descuento_Maximo_%', 
                         '%_Items_Con_Descuento', 'Es_Cazador_Descuentos']],
        on='Email', how='left'
    )
    clientes_stats['Descuento_Promedio_%'] = clientes_stats['Descuento_Promedio_%'].fillna(0)
    clientes_stats['Descuento_Maximo_%'] = clientes_stats['Descuento_Maximo_%'].fillna(0)
    clientes_stats['%_Items_Con_Descuento'] = clientes_stats['%_Items_Con_Descuento'].fillna(0)
    
    # 5. Análisis de Antigüedad de Inventario Comprado
    antiguedad_stats = ventas_df.groupby('Email Cliente').agg({
        'Días desde Última Recepción CEG': ['mean', 'median', 'min', 'max'],
        'Total Item con IVA': 'sum',
        'Cantidad Unitarias': 'sum'
    }).reset_index()
    antiguedad_stats.columns = ['Email', 'Dias_Recepcion_Promedio', 'Dias_Recepcion_Mediana',
                                'Dias_Recepcion_Min', 'Dias_Recepcion_Max',
                                'Facturacion_Antiguedad', 'Unidades_Antiguedad']
    
    antiguedad_stats['Compra_Inventario_Fresco'] = antiguedad_stats['Dias_Recepcion_Promedio'] <= 90
    antiguedad_stats['Compra_Inventario_Viejo'] = antiguedad_stats['Dias_Recepcion_Promedio'] >= 365
    antiguedad_stats['Tipo_Comprador_Inventario'] = antiguedad_stats.apply(
        lambda x: 'Compra Fresco' if x['Compra_Inventario_Fresco']
        else 'Compra Viejo' if x['Compra_Inventario_Viejo']
        else 'Mixto', axis=1
    )
    
    clientes_stats = clientes_stats.merge(
        antiguedad_stats[['Email', 'Dias_Recepcion_Promedio', 'Dias_Recepcion_Mediana',
                          'Tipo_Comprador_Inventario', 'Compra_Inventario_Fresco', 'Compra_Inventario_Viejo']],
        on='Email', how='left'
    )
    clientes_stats['Dias_Recepcion_Promedio'] = clientes_stats['Dias_Recepcion_Promedio'].fillna(0)
    clientes_stats['Dias_Recepcion_Mediana'] = clientes_stats['Dias_Recepcion_Mediana'].fillna(0)
    
    # 6. Análisis de Márgenes y Rentabilidad por Cliente
    print("   💰 Calculando márgenes y rentabilidad...")
    
    margen_stats = ventas_df.groupby('Email Cliente').agg({
        '% Margen sobre FOB': ['mean', 'median', 'min', 'max'],
        '% Margen sobre Plataforma': ['mean', 'median', 'min', 'max'],
        '%_Compra_Sobre_FOB': ['mean', 'median', 'min', 'max'],
        '%_Compra_Sobre_Plataforma': ['mean', 'median', 'min', 'max'],
        'Total Item con IVA': 'sum',
        'Cantidad Unitarias': 'sum',
        'FOB CEG': lambda x: (x * ventas_df.loc[x.index, 'Cantidad Unitarias']).sum() if len(x) > 0 else 0,
        'Base Price CEG': lambda x: (x * ventas_df.loc[x.index, 'Cantidad Unitarias']).sum() if len(x) > 0 else 0
    }).reset_index()
    
    margen_stats.columns = ['Email', 'Margen_FOB_Promedio_%', 'Margen_FOB_Mediana_%', 
                            'Margen_FOB_Min_%', 'Margen_FOB_Max_%',
                            'Margen_Plataforma_Promedio_%', 'Margen_Plataforma_Mediana_%',
                            'Margen_Plataforma_Min_%', 'Margen_Plataforma_Max_%',
                            '%_Compra_Sobre_FOB_Promedio', '%_Compra_Sobre_FOB_Mediana',
                            '%_Compra_Sobre_FOB_Min', '%_Compra_Sobre_FOB_Max',
                            '%_Compra_Sobre_Plataforma_Promedio', '%_Compra_Sobre_Plataforma_Mediana',
                            '%_Compra_Sobre_Plataforma_Min', '%_Compra_Sobre_Plataforma_Max',
                            'Facturacion_Margen', 'Unidades_Margen', 'Costo_FOB_Total', 'Costo_Plataforma_Total']
    
    # Calcular rentabilidad
    margen_stats['Ganancia_Estimada_FOB'] = margen_stats['Facturacion_Margen'] - margen_stats['Costo_FOB_Total']
    margen_stats['Ganancia_Estimada_Plataforma'] = margen_stats['Facturacion_Margen'] - margen_stats['Costo_Plataforma_Total']
    margen_stats['%_Rentabilidad_FOB'] = (
        (margen_stats['Ganancia_Estimada_FOB'] / margen_stats['Facturacion_Margen'] * 100) 
        if margen_stats['Facturacion_Margen'].sum() > 0 else 0
    )
    margen_stats['%_Rentabilidad_Plataforma'] = (
        (margen_stats['Ganancia_Estimada_Plataforma'] / margen_stats['Facturacion_Margen'] * 100)
        if margen_stats['Facturacion_Margen'].sum() > 0 else 0
    )
    
    # Clasificar oportunistas
    margen_stats['Es_Oportunista_FOB'] = (
        (margen_stats['Margen_FOB_Promedio_%'] < 50) | 
        (margen_stats['%_Compra_Sobre_FOB_Promedio'] < 100)
    )
    margen_stats['Es_Oportunista_Plataforma'] = (
        (margen_stats['Margen_Plataforma_Promedio_%'] < 20) |
        (margen_stats['%_Compra_Sobre_Plataforma_Promedio'] < 10)
    )
    margen_stats['Es_Oportunista'] = (
        margen_stats['Es_Oportunista_FOB'] | margen_stats['Es_Oportunista_Plataforma']
    )
    
    # Clasificar por tipo de comprador según márgenes
    def clasificar_comprador_margen(row):
        if row['Margen_FOB_Promedio_%'] < 30 or row['Margen_Plataforma_Promedio_%'] < 10:
            return 'Oportunista (Muy Bajo Margen)'
        elif row['Margen_FOB_Promedio_%'] < 50 or row['Margen_Plataforma_Promedio_%'] < 20:
            return 'Oportunista (Bajo Margen)'
        elif row['Margen_FOB_Promedio_%'] >= 100 and row['Margen_Plataforma_Promedio_%'] >= 30:
            return 'Premium (Alto Margen)'
        else:
            return 'Regular (Margen Estándar)'
    
    margen_stats['Tipo_Comprador_Margen'] = margen_stats.apply(clasificar_comprador_margen, axis=1)
    
    clientes_stats = clientes_stats.merge(
        margen_stats[['Email', 'Margen_FOB_Promedio_%', 'Margen_FOB_Mediana_%',
                     'Margen_Plataforma_Promedio_%', 'Margen_Plataforma_Mediana_%',
                     '%_Compra_Sobre_FOB_Promedio', '%_Compra_Sobre_Plataforma_Promedio',
                     '%_Rentabilidad_FOB', '%_Rentabilidad_Plataforma',
                     'Es_Oportunista', 'Tipo_Comprador_Margen',
                     'Ganancia_Estimada_FOB', 'Ganancia_Estimada_Plataforma']],
        on='Email', how='left'
    )
    clientes_stats['Margen_FOB_Promedio_%'] = clientes_stats['Margen_FOB_Promedio_%'].fillna(0)
    clientes_stats['Margen_Plataforma_Promedio_%'] = clientes_stats['Margen_Plataforma_Promedio_%'].fillna(0)
    clientes_stats['%_Compra_Sobre_FOB_Promedio'] = clientes_stats['%_Compra_Sobre_FOB_Promedio'].fillna(0)
    clientes_stats['%_Compra_Sobre_Plataforma_Promedio'] = clientes_stats['%_Compra_Sobre_Plataforma_Promedio'].fillna(0)
    clientes_stats['Es_Oportunista'] = clientes_stats['Es_Oportunista'].fillna(False)
    
    # Calcular 80/20
    clientes_stats = clientes_stats.sort_values('LTV', ascending=False)
    clientes_stats['LTV_Acumulado'] = clientes_stats['LTV'].cumsum()
    total_ltv = clientes_stats['LTV'].sum()
    clientes_stats['%_Facturacion'] = (clientes_stats['LTV_Acumulado'] / total_ltv * 100)
    clientes_stats['Es_80_20'] = clientes_stats['%_Facturacion'] <= 80
    
    # Segmentación RFV
    def clasificar_segmento(row):
        if row['LTV'] >= 50000 and row['Ordenes'] >= 6 and row['Dias_Desde_Ultima'] <= 90:
            return 'Champion'
        elif row['LTV'] >= 20000 and row['Ordenes'] >= 3 and row['Dias_Desde_Ultima'] <= 180:
            return 'Loyal Customer'
        elif row['Dias_Desde_Ultima'] > 180 and row['Ordenes'] >= 2:
            return 'At Risk'
        elif row['Ordenes'] == 1 and row['Dias_Desde_Ultima'] <= 90:
            return 'New Customer'
        elif row['Dias_Desde_Ultima'] > 365:
            return 'Lost Customer'
        else:
            return 'Regular'
    
    clientes_stats['Segmento_RFV'] = clientes_stats.apply(clasificar_segmento, axis=1)
    
    # Clasificación de salud
    def clasificar_salud(row):
        if row['Cliente_Sano'] and row['LTV'] >= 20000:
            return 'Muy Sano'
        elif row['Cliente_Sano']:
            return 'Sano'
        elif row['Dias_Desde_Ultima'] <= 180:
            return 'Regular'
        else:
            return 'Requiere Atención'
    
    clientes_stats['Salud_Cliente'] = clientes_stats.apply(clasificar_salud, axis=1)
    
    return clientes_stats


def generate_client_analysis_excel():
    """Genera Excel completo de análisis de clientes."""
    print("🔄 Generando Análisis Completo de Clientes...")
    
    # Cargar datos
    ventas_df = load_ventas()
    
    if len(ventas_df) == 0:
        print("   ⚠️  No hay datos de ventas para analizar")
        return
    
    # Analizar clientes
    clientes_df = analyze_clients(ventas_df)
    
    if len(clientes_df) == 0:
        print("   ⚠️  No se pudieron analizar clientes")
        return
    
    # Crear directorio de salida
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    print(f"\n💾 Creando archivo Excel: {OUTPUT_EXCEL}")
    
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        # 1. TOP 100 Clientes (combinando criterios) - CON MÉTRICAS DE MARKETING
        top_100 = clientes_df.head(100).copy()
        top_100['Rank'] = range(1, len(top_100) + 1)
        top_100_cols = ['Rank', 'Email', 'Nombre', 'Apellido', 'CUIT', 'LTV', 'Ordenes', 
                       'Ticket_Promedio', 'Dias_Desde_Ultima', 'Reincidente', 'Cliente_Sano',
                       'Segmento_RFV', 'Salud_Cliente', 'Categoria_Favorita', 'Marca_Favorita',
                       'Fan_Marca_Nombre', '%_Facturacion_Marca_Dominante', 'Es_Fan_Marca',
                       'Fiel_Vertical_Nombre', '%_Facturacion_Categoria_Dominante', 'Es_Fiel_Vertical',
                       'Diversidad_Compra', 'Marcas_Unicas', 'Categorias_Unicas',
                       'Descuento_Promedio_%', '%_Items_Con_Descuento', 'Es_Cazador_Descuentos',
                       'Tipo_Comprador_Inventario', 'Dias_Recepcion_Promedio',
                       'Margen_FOB_Promedio_%', 'Margen_Plataforma_Promedio_%',
                       '%_Compra_Sobre_FOB_Promedio', '%_Compra_Sobre_Plataforma_Promedio',
                       '%_Rentabilidad_FOB', '%_Rentabilidad_Plataforma',
                       'Es_Oportunista', 'Tipo_Comprador_Margen',
                       'Unidades_Totales', 'SKUs_Unicos', 'Primera_Compra', 'Ultima_Compra']
        # Filtrar columnas que existen
        top_100_cols = [col for col in top_100_cols if col in top_100.columns]
        top_100[top_100_cols].to_excel(writer, sheet_name='01_TOP_100_Clientes', index=False)
        auto_adjust_column_widths(writer, '01_TOP_100_Clientes', top_100[top_100_cols])
        print(f"   ✅ TOP 100 clientes generado (con métricas de marketing)")
        
        # 2. Clientes 80/20
        clientes_80_20 = clientes_df[clientes_df['Es_80_20']].copy()
        clientes_80_20['Rank'] = range(1, len(clientes_80_20) + 1)
        clientes_80_20[top_100_cols].to_excel(writer, sheet_name='02_Clientes_80_20', index=False)
        auto_adjust_column_widths(writer, '02_Clientes_80_20', clientes_80_20[top_100_cols])
        print(f"   ✅ Clientes 80/20 generado ({len(clientes_80_20)} clientes)")
        
        # 3. Clientes Reincidentes
        clientes_reincidentes = clientes_df[clientes_df['Reincidente']].copy().sort_values('LTV', ascending=False)
        clientes_reincidentes['Rank'] = range(1, len(clientes_reincidentes) + 1)
        clientes_reincidentes[top_100_cols].to_excel(writer, sheet_name='03_Clientes_Reincidentes', index=False)
        auto_adjust_column_widths(writer, '03_Clientes_Reincidentes', clientes_reincidentes[top_100_cols])
        print(f"   ✅ Clientes reincidentes generado ({len(clientes_reincidentes)} clientes)")
        
        # 4. Clientes Sanos
        clientes_sanos = clientes_df[clientes_df['Cliente_Sano']].copy().sort_values('LTV', ascending=False)
        clientes_sanos['Rank'] = range(1, len(clientes_sanos) + 1)
        clientes_sanos[top_100_cols].to_excel(writer, sheet_name='04_Clientes_Sanos', index=False)
        auto_adjust_column_widths(writer, '04_Clientes_Sanos', clientes_sanos[top_100_cols])
        print(f"   ✅ Clientes sanos generado ({len(clientes_sanos)} clientes)")
        
        # 5. Segmentación RFV
        segmentacion_rfv = clientes_df.groupby('Segmento_RFV').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Ordenes': 'sum',
            'Ticket_Promedio': 'mean'
        }).reset_index()
        segmentacion_rfv.columns = ['Segmento', 'Cantidad_Clientes', 'Facturacion_Total', 'Total_Ordenes', 'Ticket_Promedio']
        segmentacion_rfv['%_Clientes'] = (segmentacion_rfv['Cantidad_Clientes'] / len(clientes_df) * 100).round(2)
        segmentacion_rfv['%_Facturacion'] = (segmentacion_rfv['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        segmentacion_rfv.to_excel(writer, sheet_name='05_Segmentacion_RFV', index=False)
        auto_adjust_column_widths(writer, '05_Segmentacion_RFV', segmentacion_rfv)
        print(f"   ✅ Segmentación RFV generada")
        
        # 6. Análisis por Salud
        salud_analisis = clientes_df.groupby('Salud_Cliente').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Ordenes': 'sum'
        }).reset_index()
        salud_analisis.columns = ['Salud', 'Cantidad_Clientes', 'Facturacion_Total', 'Total_Ordenes']
        salud_analisis['%_Clientes'] = (salud_analisis['Cantidad_Clientes'] / len(clientes_df) * 100).round(2)
        salud_analisis['%_Facturacion'] = (salud_analisis['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        salud_analisis.to_excel(writer, sheet_name='06_Analisis_Salud', index=False)
        auto_adjust_column_widths(writer, '06_Analisis_Salud', salud_analisis)
        print(f"   ✅ Análisis de salud generado")
        
        # 7. Fans de Marcas (70%+ facturación en una marca)
        fans_marcas = clientes_df[clientes_df['Es_Fan_Marca'] == True].copy().sort_values('LTV', ascending=False)
        fans_marcas['Rank'] = range(1, len(fans_marcas) + 1)
        fans_cols = ['Rank', 'Email', 'Nombre', 'Apellido', 'CUIT', 'LTV', 'Ordenes',
                    'Fan_Marca_Nombre', 'Marca_Dominante', '%_Facturacion_Marca_Dominante',
                    'Marcas_Unicas', 'Segmento_RFV', 'Salud_Cliente']
        fans_cols = [col for col in fans_cols if col in fans_marcas.columns]
        fans_marcas[fans_cols].to_excel(writer, sheet_name='07_Fans_de_Marcas', index=False)
        auto_adjust_column_widths(writer, '07_Fans_de_Marcas', fans_marcas[fans_cols])
        print(f"   ✅ Fans de marcas generado ({len(fans_marcas)} clientes)")
        
        # 8. Fieles a Verticales (60%+ facturación en una categoría)
        fieles_verticales = clientes_df[clientes_df['Es_Fiel_Vertical'] == True].copy().sort_values('LTV', ascending=False)
        fieles_verticales['Rank'] = range(1, len(fieles_verticales) + 1)
        fieles_cols = ['Rank', 'Email', 'Nombre', 'Apellido', 'CUIT', 'LTV', 'Ordenes',
                      'Fiel_Vertical_Nombre', 'Categoria_Dominante', '%_Facturacion_Categoria_Dominante',
                      'Categorias_Unicas', 'Segmento_RFV', 'Salud_Cliente']
        fieles_cols = [col for col in fieles_cols if col in fieles_verticales.columns]
        fieles_verticales[fieles_cols].to_excel(writer, sheet_name='08_Fieles_a_Verticales', index=False)
        auto_adjust_column_widths(writer, '08_Fieles_a_Verticales', fieles_verticales[fieles_cols])
        print(f"   ✅ Fieles a verticales generado ({len(fieles_verticales)} clientes)")
        
        # 9. Análisis de Diversidad de Compra
        diversidad_analisis = clientes_df.groupby('Diversidad_Compra').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Ordenes': 'sum',
            'Ticket_Promedio': 'mean',
            'Marcas_Unicas': 'mean',
            'Categorias_Unicas': 'mean'
        }).reset_index()
        diversidad_analisis.columns = ['Diversidad', 'Cantidad_Clientes', 'Facturacion_Total', 
                                      'Total_Ordenes', 'Ticket_Promedio', 'Marcas_Promedio', 'Categorias_Promedio']
        diversidad_analisis['%_Clientes'] = (diversidad_analisis['Cantidad_Clientes'] / len(clientes_df) * 100).round(2)
        diversidad_analisis['%_Facturacion'] = (diversidad_analisis['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        diversidad_analisis.to_excel(writer, sheet_name='09_Diversidad_Compra', index=False)
        auto_adjust_column_widths(writer, '09_Diversidad_Compra', diversidad_analisis)
        print(f"   ✅ Análisis de diversidad generado")
        
        # 10. Cazadores de Descuentos
        cazadores = clientes_df[clientes_df['Es_Cazador_Descuentos'] == True].copy().sort_values('Descuento_Promedio_%', ascending=False)
        cazadores['Rank'] = range(1, len(cazadores) + 1)
        cazadores_cols = ['Rank', 'Email', 'Nombre', 'Apellido', 'CUIT', 'LTV', 'Ordenes',
                         'Descuento_Promedio_%', 'Descuento_Maximo_%', '%_Items_Con_Descuento',
                         'Ticket_Promedio', 'Segmento_RFV', 'Salud_Cliente']
        cazadores_cols = [col for col in cazadores_cols if col in cazadores.columns]
        cazadores[cazadores_cols].to_excel(writer, sheet_name='10_Cazadores_Descuentos', index=False)
        auto_adjust_column_widths(writer, '10_Cazadores_Descuentos', cazadores[cazadores_cols])
        print(f"   ✅ Cazadores de descuentos generado ({len(cazadores)} clientes)")
        
        # 11. Análisis de Antigüedad de Inventario Comprado
        antiguedad_analisis = clientes_df.groupby('Tipo_Comprador_Inventario').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Ordenes': 'sum',
            'Dias_Recepcion_Promedio': 'mean',
            'Dias_Recepcion_Mediana': 'mean'
        }).reset_index()
        antiguedad_analisis.columns = ['Tipo_Comprador', 'Cantidad_Clientes', 'Facturacion_Total',
                                       'Total_Ordenes', 'Dias_Recepcion_Promedio', 'Dias_Recepcion_Mediana']
        antiguedad_analisis['%_Clientes'] = (antiguedad_analisis['Cantidad_Clientes'] / len(clientes_df) * 100).round(2)
        antiguedad_analisis['%_Facturacion'] = (antiguedad_analisis['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        antiguedad_analisis.to_excel(writer, sheet_name='11_Antiguedad_Inventario', index=False)
        auto_adjust_column_widths(writer, '11_Antiguedad_Inventario', antiguedad_analisis)
        print(f"   ✅ Análisis de antigüedad de inventario generado")
        
        # 12. Top Fans por Marca
        top_fans_marca = fans_marcas.groupby('Fan_Marca_Nombre').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Ordenes': 'sum'
        }).reset_index()
        top_fans_marca.columns = ['Marca', 'Cantidad_Fans', 'Facturacion_Total', 'Total_Ordenes']
        top_fans_marca = top_fans_marca.sort_values('Cantidad_Fans', ascending=False)
        top_fans_marca['%_Facturacion'] = (top_fans_marca['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        top_fans_marca.to_excel(writer, sheet_name='12_Top_Fans_por_Marca', index=False)
        auto_adjust_column_widths(writer, '12_Top_Fans_por_Marca', top_fans_marca)
        print(f"   ✅ Top fans por marca generado")
        
        # 13. Top Fieles por Vertical
        top_fieles_vertical = fieles_verticales.groupby('Fiel_Vertical_Nombre').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Ordenes': 'sum'
        }).reset_index()
        top_fieles_vertical.columns = ['Vertical', 'Cantidad_Fieles', 'Facturacion_Total', 'Total_Ordenes']
        top_fieles_vertical = top_fieles_vertical.sort_values('Cantidad_Fieles', ascending=False)
        top_fieles_vertical['%_Facturacion'] = (top_fieles_vertical['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        top_fieles_vertical.to_excel(writer, sheet_name='13_Top_Fieles_por_Vertical', index=False)
        auto_adjust_column_widths(writer, '13_Top_Fieles_por_Vertical', top_fieles_vertical)
        print(f"   ✅ Top fieles por vertical generado")
        
        # 14. Análisis de Márgenes y Rentabilidad
        margen_analisis = clientes_df.groupby('Tipo_Comprador_Margen').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Ordenes': 'sum',
            'Margen_FOB_Promedio_%': 'mean',
            'Margen_Plataforma_Promedio_%': 'mean',
            'Ganancia_Estimada_FOB': 'sum',
            'Ganancia_Estimada_Plataforma': 'sum'
        }).reset_index()
        margen_analisis.columns = ['Tipo_Comprador', 'Cantidad_Clientes', 'Facturacion_Total', 
                                   'Total_Ordenes', 'Margen_FOB_Promedio', 'Margen_Plataforma_Promedio',
                                   'Ganancia_FOB_Total', 'Ganancia_Plataforma_Total']
        margen_analisis['%_Clientes'] = (margen_analisis['Cantidad_Clientes'] / len(clientes_df) * 100).round(2)
        margen_analisis['%_Facturacion'] = (margen_analisis['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        margen_analisis['%_Rentabilidad_FOB'] = (
            (margen_analisis['Ganancia_FOB_Total'] / margen_analisis['Facturacion_Total'] * 100).round(2)
        )
        margen_analisis['%_Rentabilidad_Plataforma'] = (
            (margen_analisis['Ganancia_Plataforma_Total'] / margen_analisis['Facturacion_Total'] * 100).round(2)
        )
        margen_analisis.to_excel(writer, sheet_name='14_Analisis_Margenes_Rentabilidad', index=False)
        auto_adjust_column_widths(writer, '14_Analisis_Margenes_Rentabilidad', margen_analisis)
        print(f"   ✅ Análisis de márgenes y rentabilidad generado")
        
        # 15. Oportunistas (Clientes que compran con márgenes muy bajos)
        oportunistas = clientes_df[clientes_df['Es_Oportunista'] == True].copy().sort_values('LTV', ascending=False)
        oportunistas['Rank'] = range(1, len(oportunistas) + 1)
        oportunistas_cols = ['Rank', 'Email', 'Nombre', 'Apellido', 'CUIT', 'LTV', 'Ordenes',
                           'Margen_FOB_Promedio_%', 'Margen_Plataforma_Promedio_%',
                           '%_Compra_Sobre_FOB_Promedio', '%_Compra_Sobre_Plataforma_Promedio',
                           '%_Rentabilidad_FOB', '%_Rentabilidad_Plataforma',
                           'Tipo_Comprador_Margen', 'Descuento_Promedio_%',
                           'Es_Cazador_Descuentos', 'Segmento_RFV', 'Salud_Cliente']
        oportunistas_cols = [col for col in oportunistas_cols if col in oportunistas.columns]
        oportunistas[oportunistas_cols].to_excel(writer, sheet_name='15_Oportunistas', index=False)
        auto_adjust_column_widths(writer, '15_Oportunistas', oportunistas[oportunistas_cols])
        print(f"   ✅ Oportunistas generado ({len(oportunistas)} clientes)")
        
        # 16. Correlación Volumen vs Margen (¿Los que más compran son oportunistas?)
        correlacion_volumen_margen = clientes_df.copy()
        correlacion_volumen_margen['Rank_Volumen'] = correlacion_volumen_margen['LTV'].rank(ascending=False)
        correlacion_volumen_margen['Rank_Margen'] = correlacion_volumen_margen['Margen_FOB_Promedio_%'].rank(ascending=False)
        correlacion_volumen_margen['Diferencia_Rank'] = abs(correlacion_volumen_margen['Rank_Volumen'] - correlacion_volumen_margen['Rank_Margen'])
        
        # Clasificar relación volumen-margen
        def clasificar_volumen_margen(row):
            if row['Rank_Volumen'] <= 100 and row['Margen_FOB_Promedio_%'] < 50:
                return 'Alto Volumen - Bajo Margen (Oportunista)'
            elif row['Rank_Volumen'] <= 100 and row['Margen_FOB_Promedio_%'] >= 100:
                return 'Alto Volumen - Alto Margen (Ideal)'
            elif row['Rank_Volumen'] > 500 and row['Margen_FOB_Promedio_%'] >= 100:
                return 'Bajo Volumen - Alto Margen (Potencial)'
            else:
                return 'Regular'
        
        correlacion_volumen_margen['Relacion_Volumen_Margen'] = correlacion_volumen_margen.apply(clasificar_volumen_margen, axis=1)
        
        relacion_analisis = correlacion_volumen_margen.groupby('Relacion_Volumen_Margen').agg({
            'Email': 'count',
            'LTV': 'sum',
            'Margen_FOB_Promedio_%': 'mean',
            'Margen_Plataforma_Promedio_%': 'mean'
        }).reset_index()
        relacion_analisis.columns = ['Relacion', 'Cantidad_Clientes', 'Facturacion_Total',
                                     'Margen_FOB_Promedio', 'Margen_Plataforma_Promedio']
        relacion_analisis['%_Clientes'] = (relacion_analisis['Cantidad_Clientes'] / len(clientes_df) * 100).round(2)
        relacion_analisis['%_Facturacion'] = (relacion_analisis['Facturacion_Total'] / clientes_df['LTV'].sum() * 100).round(2)
        relacion_analisis.to_excel(writer, sheet_name='16_Correlacion_Volumen_Margen', index=False)
        auto_adjust_column_widths(writer, '16_Correlacion_Volumen_Margen', relacion_analisis)
        print(f"   ✅ Correlación volumen-margen generada")
        
        # 17. Todos los Clientes (completo) - CON TODAS LAS MÉTRICAS
        todos_clientes = clientes_df.copy()
        todos_clientes['Rank'] = range(1, len(todos_clientes) + 1)
        todas_cols = ['Rank', 'Email', 'Nombre', 'Apellido', 'CUIT', 'LTV', 'Ordenes', 
                     'Ticket_Promedio', 'Dias_Desde_Ultima', 'Reincidente', 'Cliente_Sano',
                     'Segmento_RFV', 'Salud_Cliente', 'Categoria_Favorita', 'Marca_Favorita',
                     'Fan_Marca_Nombre', '%_Facturacion_Marca_Dominante', 'Es_Fan_Marca',
                     'Fiel_Vertical_Nombre', '%_Facturacion_Categoria_Dominante', 'Es_Fiel_Vertical',
                     'Diversidad_Compra', 'Marcas_Unicas', 'Categorias_Unicas',
                     'Descuento_Promedio_%', '%_Items_Con_Descuento', 'Es_Cazador_Descuentos',
                     'Tipo_Comprador_Inventario', 'Dias_Recepcion_Promedio', 'Dias_Recepcion_Mediana',
                     'Margen_FOB_Promedio_%', 'Margen_Plataforma_Promedio_%',
                     '%_Compra_Sobre_FOB_Promedio', '%_Compra_Sobre_Plataforma_Promedio',
                     '%_Rentabilidad_FOB', '%_Rentabilidad_Plataforma',
                     'Es_Oportunista', 'Tipo_Comprador_Margen',
                     'Unidades_Totales', 'SKUs_Unicos', 'Primera_Compra', 'Ultima_Compra']
        todas_cols = [col for col in todas_cols if col in todos_clientes.columns]
        todos_clientes[todas_cols].to_excel(writer, sheet_name='17_Todos_Los_Clientes', index=False)
        auto_adjust_column_widths(writer, '17_Todos_Los_Clientes', todos_clientes[todas_cols])
        print(f"   ✅ Todos los clientes generado ({len(todos_clientes)} clientes con todas las métricas)")
        
        # 8. Resumen Ejecutivo
        resumen_data = {
            'Métrica': [
                'Total Clientes',
                'Clientes Reincidentes',
                'Clientes Sanos',
                'Clientes 80/20',
                'Facturación Total',
                'Facturación 80/20',
                'Ticket Promedio',
                'Órdenes Promedio',
                'Tasa Reincidencia',
                'LTV Promedio'
            ],
            'Valor': [
                len(clientes_df),
                clientes_df['Reincidente'].sum(),
                clientes_df['Cliente_Sano'].sum(),
                clientes_df['Es_80_20'].sum(),
                f"${clientes_df['LTV'].sum():,.2f} USD",
                f"${clientes_df[clientes_df['Es_80_20']]['LTV'].sum():,.2f} USD",
                f"${clientes_df['Ticket_Promedio'].mean():,.2f} USD",
                f"{clientes_df['Ordenes'].mean():.2f}",
                f"{(clientes_df['Reincidente'].sum() / len(clientes_df) * 100):.1f}%",
                f"${clientes_df['LTV'].mean():,.2f} USD"
            ]
        }
        resumen_df = pd.DataFrame(resumen_data)
        resumen_df.to_excel(writer, sheet_name='00_Resumen_Ejecutivo', index=False)
        auto_adjust_column_widths(writer, '00_Resumen_Ejecutivo', resumen_df)
        print(f"   ✅ Resumen ejecutivo generado")
    
    print(f"\n   ✅ Archivo Excel generado: {OUTPUT_EXCEL}")
    print(f"\n📋 Hojas creadas:")
    print(f"   00. Resumen Ejecutivo")
    print(f"   01. TOP 100 Clientes (con métricas de marketing)")
    print(f"   02. Clientes 80/20")
    print(f"   03. Clientes Reincidentes")
    print(f"   04. Clientes Sanos")
    print(f"   05. Segmentación RFV")
    print(f"   06. Análisis de Salud")
    print(f"   07. Fans de Marcas")
    print(f"   08. Fieles a Verticales")
    print(f"   09. Diversidad de Compra")
    print(f"   10. Cazadores de Descuentos")
    print(f"   11. Antigüedad de Inventario Comprado")
    print(f"   12. Top Fans por Marca")
    print(f"   13. Top Fieles por Vertical")
    print(f"   14. Análisis de Márgenes y Rentabilidad")
    print(f"   15. Oportunistas (Bajo Margen)")
    print(f"   16. Correlación Volumen vs Margen")
    print(f"   17. Todos Los Clientes (completo con todas las métricas)")


if __name__ == "__main__":
    if not HAS_PANDAS:
        print("Instalando pandas...")
        import subprocess
        subprocess.check_call(["pip3", "install", "pandas", "openpyxl", "--break-system-packages"])
        import pandas as pd
        from openpyxl import load_workbook
    
    generate_client_analysis_excel()
    print("\n✨ Proceso completado!")
