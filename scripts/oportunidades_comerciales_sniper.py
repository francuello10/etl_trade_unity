#!/usr/bin/env python3
"""
Oportunidades Comerciales Tipo Sniper - Trade Unity

ENFOQUE: VENDER EL STOCK ACTUAL (no reposición)

Análisis para identificar oportunidades de venta puntual del stock disponible:
1. Clientes que barrieron un producto y hay productos relacionados/familia en stock actual
2. Clientes que compraron fuerte un producto y hay productos relacionados/complementarios en stock (upselling)
3. Clientes que barrieron stock de un producto específico (muestra capacidad de compra)
4. Oportunidades de precio - clientes que compraron más barato que precio actual publicado
5. Recompra (menor importancia) - clientes que compraron fuerte y hay stock nuevo

Genera Excel con oportunidades comerciales accionables para el equipo de ventas.
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import os

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Archivos
VENTAS_CSV = "inputs/ventas_historicas_items.csv"
STOCK_ERP = "fuentes/stock_erp.csv"
CATALOGO_TU = "fuentes/catalogo_trade_unity.csv"
PUBLICACIONES_CSV = "fuentes/publicaciones_productos.csv"
OUTPUT_DIR = "outputs"
OUTPUT_EXCEL = f"{OUTPUT_DIR}/TradeUnity_Sniper_Commercial_Opportunities.xlsx"


def parse_decimal(value: str) -> float:
    """Convierte string a float."""
    if not value or value == "" or pd.isna(value):
        return 0.0
    
    value_str = str(value).strip().replace("$", "").replace(" ", "").replace("%", "")
    value_str = value_str.replace(",", ".")
    
    try:
        return float(value_str)
    except (ValueError, TypeError):
        return 0.0


def auto_adjust_column_widths(writer, sheet_name, df):
    """Ajusta automáticamente el ancho de las columnas en Excel."""
    try:
        worksheet = writer.sheets[sheet_name]
        
        for idx, col in enumerate(df.columns, 1):
            column_letter = worksheet.cell(row=1, column=idx).column_letter
            try:
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                )
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            except:
                # Si hay problemas con una columna, usar ancho por defecto
                worksheet.column_dimensions[column_letter].width = 15
    except Exception as e:
        print(f"   ⚠️  No se pudo ajustar columnas en {sheet_name}: {e}")


def load_ventas():
    """Carga datos de ventas."""
    print("📖 Cargando datos de ventas...")
    
    if not os.path.exists(VENTAS_CSV):
        print(f"   ⚠️  Archivo de ventas no encontrado: {VENTAS_CSV}")
        return pd.DataFrame()
    
    df = pd.read_csv(VENTAS_CSV, encoding='utf-8-sig')
    
    # Convertir fechas
    if 'Fecha Creación' in df.columns:
        df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación'], errors='coerce')
    
    # Convertir numéricos
    numeric_cols = [
        'Cantidad Unitarias', 'Total Item con IVA', 'Precio Venta Unitario',
        'Precio Original', 'Precio Venta', 'FOB CEG', 'Base Price CEG'
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '.').str.replace('$', '').str.strip()
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    print(f"   ✅ {len(df)} registros de ventas cargados")
    return df


def load_stock():
    """Carga stock actual del ERP."""
    print("📖 Cargando stock del ERP...")
    
    if not os.path.exists(STOCK_ERP):
        print(f"   ⚠️  Archivo de stock no encontrado: {STOCK_ERP}")
        return pd.DataFrame()
    
    df = pd.read_csv(STOCK_ERP, encoding='utf-8-sig')
    
    # Limpiar y convertir
    df['Pronosticado con pendiente'] = pd.to_numeric(
        df['Pronosticado con pendiente'].astype(str).str.replace(',', '.'), 
        errors='coerce'
    ).fillna(0)
    
    df['Box Qty'] = pd.to_numeric(
        df['Box Qty'].astype(str).str.replace(',', '.'), 
        errors='coerce'
    ).fillna(1)
    
    # Calcular unidades
    df['Stock Unidades'] = df['Pronosticado con pendiente'] * df['Box Qty']
    
    print(f"   ✅ {len(df)} productos con stock cargados")
    return df


def load_catalog():
    """Carga catálogo TU para obtener rubro y marca."""
    print("📖 Cargando catálogo TU...")
    
    if not os.path.exists(CATALOGO_TU):
        print(f"   ⚠️  Archivo de catálogo no encontrado: {CATALOGO_TU}")
        return pd.DataFrame()
    
    df = pd.read_csv(CATALOGO_TU, encoding='utf-8-sig')
    
    print(f"   ✅ {len(df)} productos en catálogo")
    return df


def load_precios_actuales():
    """Carga precios actuales (liquidación enero/febrero 2026)."""
    print("📖 Cargando precios actuales (liquidación enero/febrero 2026)...")
    
    if not os.path.exists(PUBLICACIONES_CSV):
        print(f"   ⚠️  Archivo de publicaciones no encontrado: {PUBLICACIONES_CSV}")
        return pd.DataFrame()
    
    df = pd.read_csv(PUBLICACIONES_CSV, encoding='utf-8-sig')
    
    # Obtener precio actual (última columna: liquidación enero/febrero 2026)
    precio_col = 'Precio LIQUIDACION ENERO/FEBRERO 2026  unitario neto'
    
    if precio_col in df.columns:
        df['Precio Actual Publicado'] = pd.to_numeric(
            df[precio_col].astype(str).str.replace(',', '.').str.replace('$', '').str.strip(),
            errors='coerce'
        ).fillna(0)
    else:
        # Si no existe, buscar la última columna de precio
        precio_cols = [col for col in df.columns if 'Precio' in col and 'unitario' in col.lower()]
        if precio_cols:
            ultima_col = precio_cols[-1]
            df['Precio Actual Publicado'] = pd.to_numeric(
                df[ultima_col].astype(str).str.replace(',', '.').str.replace('$', '').str.strip(),
                errors='coerce'
            ).fillna(0)
        else:
            df['Precio Actual Publicado'] = 0
    
    print(f"   ✅ Precios actuales cargados")
    return df[['sku', 'Precio Actual Publicado']]


def obtener_productos_relacionados(sku, catalog_df_indexed, stock_df, ventas_df):
    """
    Identifica productos relacionados por:
    - Misma marca
    - Misma categoría (2° nivel)
    """
    try:
        if sku not in catalog_df_indexed.index:
            return []
        
        producto = catalog_df_indexed.loc[sku]
        
        # Obtener marca y categoría
        marca = str(producto.get('Marca', '') or producto.get('Brand Name CEG', '')).strip()
        categoria_2 = str(producto.get('Categoría (2° Nivel)', '')).strip()
        
        if not marca and not categoria_2:
            return []
        
        # Filtrar productos relacionados (excluyendo el mismo SKU)
        mask = (catalog_df_indexed.index != sku)
        if marca:
            mask = mask & (catalog_df_indexed['Marca'].fillna('').astype(str).str.strip() == marca)
        if categoria_2:
            mask = mask & (catalog_df_indexed['Categoría (2° Nivel)'].fillna('').astype(str).str.strip() == categoria_2)
        
        relacionados = catalog_df_indexed[mask].copy()
        
        if len(relacionados) == 0:
            return []
        
        # Obtener stock disponible de productos relacionados
        stock_map = dict(zip(
            stock_df['D365 Reference'].str.upper().str.strip(),
            stock_df['Stock Unidades']
        ))
        
        # Mapear SKU a D365 Reference desde catálogo original
        catalog_original = catalog_df_indexed.reset_index()
        sku_to_d365 = dict(zip(
            catalog_original['sku'].str.upper().str.strip(),
            catalog_original['Código de Producto (D365)'].fillna('').astype(str).str.upper().str.strip()
        ))
        
        productos_con_stock = []
        for sku_rel in relacionados.index:
            d365_ref = sku_to_d365.get(str(sku_rel).upper().strip(), '')
            stock_disponible = stock_map.get(d365_ref, 0)
            
            if stock_disponible > 0:
                productos_con_stock.append({
                    'SKU': sku_rel,
                    'Nombre Producto': str(relacionados.loc[sku_rel, 'Nombre del Producto']),
                    'Stock Disponible': stock_disponible,
                    'Marca': marca,
                    'Categoría': categoria_2
                })
        
        return productos_con_stock
    except Exception as e:
        print(f"   ⚠️  Error obteniendo productos relacionados para {sku}: {e}")
        return []


def analizar_productos_relacionados_stock(ventas_df, stock_df, catalog_df):
    """
    PRIORIDAD 1: Clientes que barrieron un producto y hay productos relacionados/familia en stock actual.
    Oportunidad de vender productos relacionados del stock disponible.
    """
    print("🎯 Analizando clientes que barrieron productos relacionados (stock disponible)...")
    
    # Preparar catálogo con índice SKU
    catalog_df_indexed = catalog_df.set_index('sku')
    
    # Obtener SKUs con stock disponible
    stock_disponible = stock_df[stock_df['Stock Unidades'] > 0].copy()
    skus_con_stock = set(stock_df['D365 Reference'].str.upper().str.strip())
    
    # Mapear D365 a SKU
    d365_to_sku = dict(zip(
        catalog_df['Código de Producto (D365)'].str.upper().str.strip(),
        catalog_df['sku'].str.upper().str.strip()
    ))
    
    # Identificar clientes que barrieron stock (compraron significativamente)
    ventas_df['SKU_Upper'] = ventas_df['SKU'].str.upper().str.strip()
    
    # Agrupar por cliente y SKU para identificar compras significativas
    compras_por_cliente_sku = ventas_df.groupby(['Email Cliente', 'SKU']).agg({
        'Cantidad Unitarias': 'sum',
        'Total Item con IVA': 'sum',
        'Fecha Creación': 'max',
        'Número de Orden': 'nunique',
        'Precio Venta Unitario': 'mean',
        'Nombre Cliente': 'first',
        'Categoría (2° Nivel)': 'first',
        'Brand Name CEG': 'first',
        'Nombre Producto': 'first'
    }).reset_index()
    
    # Renombrar columna para consistencia
    compras_por_cliente_sku = compras_por_cliente_sku.rename(columns={'Número de Orden': 'Número de Órdenes'})
    
    # Calcular umbrales por SKU
    umbrales_por_sku = ventas_df.groupby('SKU')['Cantidad Unitarias'].quantile(0.75)
    
    # Filtrar compras significativas (percentil 75 o mínimo 50 unidades)
    compras_significativas = compras_por_cliente_sku[
        compras_por_cliente_sku.apply(
            lambda row: row['Cantidad Unitarias'] >= max(umbrales_por_sku.get(row['SKU'], 0), 50),
            axis=1
        )
    ]
    
    oportunidades = []
    
    for _, compra in compras_significativas.iterrows():
        sku_comprado = compra['SKU'].upper().strip()
        
        # Obtener productos relacionados con stock disponible
        productos_relacionados = obtener_productos_relacionados(
            sku_comprado, 
            catalog_df_indexed, 
            stock_df, 
            ventas_df
        )
        
        if len(productos_relacionados) > 0:
            # Calcular stock total disponible de productos relacionados
            stock_total_relacionados = sum(p['Stock Disponible'] for p in productos_relacionados)
            
            # Obtener capacidad de compra del cliente
            capacidad_compra = compra['Cantidad Unitarias'] / compra['Número de Órdenes'] if compra['Número de Órdenes'] > 0 else compra['Cantidad Unitarias']
            
            oportunidades.append({
                'Email Cliente': compra['Email Cliente'],
                'Nombre Cliente': compra['Nombre Cliente'],
                'SKU Comprado (Barrió)': sku_comprado,
                'Producto Comprado': compra['Nombre Producto'],
                'Unidades Compradas (Histórico)': compra['Cantidad Unitarias'],
                'Capacidad de Compra Promedio': capacidad_compra,
                'Productos Relacionados Disponibles': len(productos_relacionados),
                'Stock Total Relacionados': stock_total_relacionados,
                'Marca': compra['Brand Name CEG'],
                'Categoría': compra['Categoría (2° Nivel)'],
                'Total Facturación Histórica (USD)': compra['Total Item con IVA'],
                'Última Compra': compra['Fecha Creación'],
                'Días desde Última Compra': (datetime.now() - compra['Fecha Creación']).days if pd.notna(compra['Fecha Creación']) else None,
                'Tipo Oportunidad': 'Productos Relacionados - Stock Disponible',
                'Prioridad': 'ALTA',
                'Mensaje Comercial': f"Cliente barrió {compra['Nombre Producto']}. Hay {len(productos_relacionados)} productos relacionados de {compra['Brand Name CEG']} en stock ({int(stock_total_relacionados)} unidades). Oportunidad de upselling/familia."
            })
    
    df_oportunidades = pd.DataFrame(oportunidades)
    
    if len(df_oportunidades) > 0:
        df_oportunidades = df_oportunidades.sort_values('Stock Total Relacionados', ascending=False)
        print(f"   ✅ {len(df_oportunidades)} oportunidades de productos relacionados identificadas")
    else:
        print("   ⚠️  No se encontraron oportunidades de productos relacionados")
    
    return df_oportunidades


def analizar_upselling_stock_actual(ventas_df, stock_df, catalog_df):
    """
    PRIORIDAD 2: Clientes que compraron fuerte un producto y hay productos relacionados/complementarios en stock.
    Oportunidad de upselling con stock disponible.
    """
    print("🎯 Analizando oportunidades de upselling (stock disponible)...")
    
    # Similar a productos relacionados pero enfocado en complementarios/upselling
    # Por ahora usar misma lógica pero con enfoque diferente
    return analizar_productos_relacionados_stock(ventas_df, stock_df, catalog_df)


def analizar_clientes_barren_stock(ventas_df, stock_df, catalog_df):
    """
    PRIORIDAD 3: Clientes que barrieron stock de un producto específico (ya no está en stock).
    Muestra capacidad de compra, útil para identificar clientes con capacidad de limpiar stocks.
    """
    print("🎯 Analizando clientes que barrieron stock (capacidad de compra)...")
    
    # Obtener SKUs en stock 0
    stock_cero = stock_df[stock_df['Stock Unidades'] == 0].copy()
    skus_stock_cero = set(stock_cero['D365 Reference'].str.upper().str.strip())
    
    # Agregar SKU desde catálogo si no está en stock
    if 'sku' in catalog_df.columns and 'Código de Producto (D365)' in catalog_df.columns:
        catalog_sku_map = dict(zip(
            catalog_df['Código de Producto (D365)'].str.upper().str.strip(),
            catalog_df['sku'].str.upper().str.strip()
        ))
        skus_adicionales = set()
        for d365_ref in list(skus_stock_cero):  # Convertir a lista para evitar modificar durante iteración
            if d365_ref in catalog_sku_map:
                skus_adicionales.add(catalog_sku_map[d365_ref])
        skus_stock_cero.update(skus_adicionales)
    
    # Filtrar ventas de SKUs que ahora están en stock 0
    ventas_df['SKU_Upper'] = ventas_df['SKU'].str.upper().str.strip()
    ventas_stock_cero = ventas_df[ventas_df['SKU_Upper'].isin(skus_stock_cero)].copy()
    
    if len(ventas_stock_cero) == 0:
        print("   ⚠️  No se encontraron ventas de SKUs en stock 0")
        return pd.DataFrame()
    
    # Agrupar por cliente y SKU para identificar compras significativas
    # "Significativo" = percentil 75 de unidades compradas por SKU
    umbral_significativo = ventas_stock_cero.groupby('SKU')['Cantidad Unitarias'].quantile(0.75)
    
    # También calcular mediana para identificar compradores fuertes
    mediana_por_sku = ventas_stock_cero.groupby('SKU')['Cantidad Unitarias'].median()
    
    oportunidades = []
    
    for (email, sku), group in ventas_stock_cero.groupby(['Email Cliente', 'SKU']):
        total_unidades = group['Cantidad Unitarias'].sum()
        total_facturacion = group['Total Item con IVA'].sum()
        ultima_compra = group['Fecha Creación'].max()
        num_ordenes = group['Número de Orden'].nunique()
        precio_promedio = group['Precio Venta Unitario'].mean()
        
        # Obtener umbral para este SKU
        umbral = umbral_significativo.get(sku, 0)
        mediana = mediana_por_sku.get(sku, 0)
        
        # Si compró más que el umbral O más que la mediana x 2, es una oportunidad
        # Mínimo 50 unidades para considerar "significativo"
        if total_unidades >= max(umbral, mediana * 2, 50):
            # Obtener info del producto
            producto_info = group.iloc[0]
            
            # Calcular capacidad de compra (unidades promedio por orden)
            capacidad_compra = total_unidades / num_ordenes if num_ordenes > 0 else total_unidades
            
            oportunidades.append({
                'Email Cliente': email,
                'Nombre Cliente': producto_info.get('Nombre Cliente', ''),
                'SKU': sku,
                'Nombre Producto': producto_info.get('Nombre Producto', ''),
                'Categoría (2° Nivel)': producto_info.get('Categoría (2° Nivel)', ''),
                'Brand Name CEG': producto_info.get('Brand Name CEG', ''),
                'Total Unidades Compradas': total_unidades,
                'Unidades Promedio por Orden': capacidad_compra,
                'Total Facturación Histórica (USD)': total_facturacion,
                'Precio Promedio Histórico (USD)': precio_promedio,
                'Número de Órdenes': num_ordenes,
                'Última Compra': ultima_compra,
                'Días desde Última Compra': (datetime.now() - ultima_compra).days if pd.notna(ultima_compra) else None,
                'Stock Actual': 0,
                'Tipo Oportunidad': 'Barre Stock',
                'Prioridad': 'ALTA' if total_unidades >= 200 else 'MEDIA',
                'Mensaje Comercial': f"Cliente compró {int(total_unidades)} unidades. SKU ahora en stock 0. Oportunidad de liquidación o reposición."
            })
    
    df_oportunidades = pd.DataFrame(oportunidades)
    
    if len(df_oportunidades) > 0:
        # Ordenar por unidades compradas (descendente)
        df_oportunidades = df_oportunidades.sort_values('Total Unidades Compradas', ascending=False)
        print(f"   ✅ {len(df_oportunidades)} oportunidades de clientes que barren stock identificadas")
    else:
        print("   ⚠️  No se encontraron oportunidades de clientes que barren stock")
    
    return df_oportunidades


def analizar_stock_nuevo_recompra(ventas_df, stock_df, catalog_df):
    """
    PRIORIDAD 5 (MENOR IMPORTANCIA): Clientes que compraron fuerte un SKU y ahora hay stock nuevo disponible.
    Oportunidad de recompra - mantener pero con menor importancia.
    """
    print("🎯 Analizando oportunidades de recompra (stock nuevo) - MENOR PRIORIDAD...")
    
    # Obtener SKUs con stock disponible (>0)
    stock_disponible = stock_df[stock_df['Stock Unidades'] > 0].copy()
    skus_con_stock = set(stock_disponible['D365 Reference'].str.upper().str.strip())
    
    # Agregar SKU desde catálogo
    if 'sku' in catalog_df.columns and 'Código de Producto (D365)' in catalog_df.columns:
        catalog_sku_map = dict(zip(
            catalog_df['Código de Producto (D365)'].str.upper().str.strip(),
            catalog_df['sku'].str.upper().str.strip()
        ))
        skus_adicionales = set()
        for d365_ref in list(skus_con_stock):  # Convertir a lista para evitar modificar durante iteración
            if d365_ref in catalog_sku_map:
                skus_adicionales.add(catalog_sku_map[d365_ref])
        skus_con_stock.update(skus_adicionales)
    
    # Filtrar ventas de SKUs que ahora tienen stock
    ventas_df['SKU_Upper'] = ventas_df['SKU'].str.upper().str.strip()
    ventas_con_stock = ventas_df[ventas_df['SKU_Upper'].isin(skus_con_stock)].copy()
    
    if len(ventas_con_stock) == 0:
        print("   ⚠️  No se encontraron ventas de SKUs con stock disponible")
        return pd.DataFrame()
    
    # Agrupar por cliente y SKU
    # "Compró fuerte" = percentil 75 de unidades compradas por SKU
    umbral_fuerte = ventas_con_stock.groupby('SKU')['Cantidad Unitarias'].quantile(0.75)
    mediana_por_sku = ventas_con_stock.groupby('SKU')['Cantidad Unitarias'].median()
    
    # Merge con stock actual
    stock_map = dict(zip(
        stock_disponible['D365 Reference'].str.upper().str.strip(),
        stock_disponible['Stock Unidades']
    ))
    
    oportunidades = []
    
    for (email, sku), group in ventas_con_stock.groupby(['Email Cliente', 'SKU']):
        total_unidades = group['Cantidad Unitarias'].sum()
        total_facturacion = group['Total Item con IVA'].sum()
        ultima_compra = group['Fecha Creación'].max()
        num_ordenes = group['Número de Orden'].nunique()
        precio_promedio = group['Precio Venta Unitario'].mean()
        
        # Obtener umbral para este SKU
        umbral = umbral_fuerte.get(sku, 0)
        mediana = mediana_por_sku.get(sku, 0)
        
        # Si compró más que el umbral O más que la mediana x 2, es una oportunidad
        if total_unidades >= max(umbral, mediana * 2, 50):  # Mínimo 50 unidades
            # Obtener stock actual
            stock_actual = stock_map.get(sku, 0)
            
            # Obtener info del producto
            producto_info = group.iloc[0]
            
            # Calcular capacidad de compra
            capacidad_compra = total_unidades / num_ordenes if num_ordenes > 0 else total_unidades
            
            # Potencial de venta (mínimo entre stock disponible y capacidad histórica)
            potencial_venta = min(stock_actual, capacidad_compra * 1.2)  # 20% más que promedio histórico
            
            oportunidades.append({
                'Email Cliente': email,
                'Nombre Cliente': producto_info.get('Nombre Cliente', ''),
                'SKU': sku,
                'Nombre Producto': producto_info.get('Nombre Producto', ''),
                'Categoría (2° Nivel)': producto_info.get('Categoría (2° Nivel)', ''),
                'Brand Name CEG': producto_info.get('Brand Name CEG', ''),
                'Total Unidades Compradas Histórico': total_unidades,
                'Unidades Promedio por Orden': capacidad_compra,
                'Stock Actual Disponible': stock_actual,
                'Potencial de Venta (Unidades)': potencial_venta,
                'Total Facturación Histórica (USD)': total_facturacion,
                'Precio Promedio Histórico (USD)': precio_promedio,
                'Número de Órdenes': num_ordenes,
                'Última Compra': ultima_compra,
                'Días desde Última Compra': (datetime.now() - ultima_compra).days if pd.notna(ultima_compra) else None,
                'Tipo Oportunidad': 'Recompra (Stock Nuevo)',
                'Prioridad': 'BAJA',  # Menor prioridad - enfoque en stock actual, no reposición
                'Mensaje Comercial': f"Cliente compró {int(total_unidades)} unidades históricamente. Stock nuevo disponible: {int(stock_actual)} unidades. Oportunidad de recompra."
            })
    
    df_oportunidades = pd.DataFrame(oportunidades)
    
    if len(df_oportunidades) > 0:
        # Ordenar por unidades compradas (descendente)
        df_oportunidades = df_oportunidades.sort_values('Total Unidades Compradas Histórico', ascending=False)
        print(f"   ✅ {len(df_oportunidades)} oportunidades de recompra identificadas")
    else:
        print("   ⚠️  No se encontraron oportunidades de recompra")
    
    return df_oportunidades


def analizar_oportunidades_precio(ventas_df, precios_df):
    """
    Identifica clientes que compraron un producto a precio más barato que el precio actual publicado.
    Oportunidad de venta con precio atractivo.
    """
    print("🎯 Analizando oportunidades de precio...")
    
    # Merge ventas con precios actuales
    ventas_df['SKU_Upper'] = ventas_df['SKU'].str.upper().str.strip()
    precios_df['sku_upper'] = precios_df['sku'].str.upper().str.strip()
    
    ventas_con_precio = ventas_df.merge(
        precios_df[['sku_upper', 'Precio Actual Publicado']],
        left_on='SKU_Upper',
        right_on='sku_upper',
        how='inner'
    )
    
    # Filtrar solo productos con precio actual > 0
    ventas_con_precio = ventas_con_precio[ventas_con_precio['Precio Actual Publicado'] > 0]
    
    if len(ventas_con_precio) == 0:
        print("   ⚠️  No se encontraron ventas con precios actuales")
        return pd.DataFrame()
    
    # Identificar compras donde precio histórico < precio actual
    # Esto significa que compró más barato, ahora puede comprar al precio actual (más caro) o negociar
    ventas_con_precio['Precio Histórico Unitario'] = ventas_con_precio['Precio Venta Unitario']
    ventas_con_precio['Diferencia Precio'] = ventas_con_precio['Precio Actual Publicado'] - ventas_con_precio['Precio Histórico Unitario']
    ventas_con_precio['% Diferencia'] = (ventas_con_precio['Diferencia Precio'] / ventas_con_precio['Precio Histórico Unitario'] * 100).fillna(0)
    
    # Filtrar: compró más barato que precio actual (diferencia > 0)
    oportunidades_precio = ventas_con_precio[ventas_con_precio['Diferencia Precio'] > 0].copy()
    
    if len(oportunidades_precio) == 0:
        print("   ⚠️  No se encontraron oportunidades de precio")
        return pd.DataFrame()
    
    # Agrupar por cliente y SKU
    oportunidades = []
    
    for (email, sku), group in oportunidades_precio.groupby(['Email Cliente', 'SKU']):
        total_unidades = group['Cantidad Unitarias'].sum()
        total_facturacion = group['Total Item con IVA'].sum()
        ultima_compra = group['Fecha Creación'].max()
        precio_historico_promedio = group['Precio Histórico Unitario'].mean()
        precio_actual = group['Precio Actual Publicado'].iloc[0]
        diferencia_pct = group['% Diferencia'].mean()
        
        # Obtener info del producto
        producto_info = group.iloc[0]
        
        # Calcular ahorro potencial si compra al precio histórico
        ahorro_potencial = (precio_actual - precio_historico_promedio) * total_unidades
        
        oportunidades.append({
            'Email Cliente': email,
            'Nombre Cliente': producto_info.get('Nombre Cliente', ''),
            'SKU': sku,
            'Nombre Producto': producto_info.get('Nombre Producto', ''),
            'Categoría (2° Nivel)': producto_info.get('Categoría (2° Nivel)', ''),
            'Brand Name CEG': producto_info.get('Brand Name CEG', ''),
            'Total Unidades Compradas': total_unidades,
            'Precio Histórico Promedio (USD)': precio_historico_promedio,
            'Precio Actual Publicado (USD)': precio_actual,
            'Diferencia Precio (USD)': precio_actual - precio_historico_promedio,
            '% Diferencia Precio': diferencia_pct,
            'Ahorro Potencial si Precio Histórico (USD)': ahorro_potencial,
            'Total Facturación Histórica (USD)': total_facturacion,
            'Última Compra': ultima_compra,
            'Días desde Última Compra': (datetime.now() - ultima_compra).days if pd.notna(ultima_compra) else None,
            'Tipo Oportunidad': 'Oportunidad Precio',
            'Prioridad': 'ALTA' if diferencia_pct > 20 else 'MEDIA',
            'Mensaje Comercial': f"Cliente compró a ${precio_historico_promedio:.2f} USD. Precio actual: ${precio_actual:.2f} USD ({diferencia_pct:.1f}% más caro). Oportunidad de negociación o mantener precio histórico."
        })
    
    df_oportunidades = pd.DataFrame(oportunidades)
    
    if len(df_oportunidades) > 0:
        # Ordenar por diferencia de precio (descendente)
        df_oportunidades = df_oportunidades.sort_values('% Diferencia Precio', ascending=False)
        print(f"   ✅ {len(df_oportunidades)} oportunidades de precio identificadas")
    else:
        print("   ⚠️  No se encontraron oportunidades de precio")
    
    return df_oportunidades


def generar_resumen_por_rubro_marca(df_oportunidades):
    """Genera resumen de oportunidades por rubro y marca."""
    if len(df_oportunidades) == 0:
        return pd.DataFrame()
    
    # Identificar columna de unidades (puede tener nombres diferentes)
    unidades_col = None
    for col in ['Total Unidades Compradas', 'Total Unidades Compradas Histórico']:
        if col in df_oportunidades.columns:
            unidades_col = col
            break
    
    if unidades_col is None:
        print("   ⚠️  No se encontró columna de unidades")
        return pd.DataFrame()
    
    resumen = df_oportunidades.groupby(['Categoría (2° Nivel)', 'Brand Name CEG']).agg({
        'Email Cliente': 'nunique',
        'SKU': 'nunique',
        unidades_col: 'sum',
        'Total Facturación Histórica (USD)': 'sum'
    }).reset_index()
    
    resumen.columns = [
        'Categoría (2° Nivel)',
        'Brand Name CEG',
        'Clientes Únicos',
        'SKUs Únicos',
        'Total Unidades',
        'Total Facturación (USD)'
    ]
    
    resumen = resumen.sort_values('Total Facturación (USD)', ascending=False)
    
    return resumen


def generate_sniper_report():
    """Genera reporte completo de oportunidades comerciales tipo sniper."""
    print("🚀 Generando Reporte de Oportunidades Comerciales Sniper...")
    print("=" * 70)
    
    # Cargar datos
    ventas_df = load_ventas()
    stock_df = load_stock()
    catalog_df = load_catalog()
    precios_df = load_precios_actuales()
    
    if len(ventas_df) == 0:
        print("❌ No hay datos de ventas. Abortando.")
        return
    
    # PRIORIDAD 1: Productos relacionados con stock disponible
    oportunidades_relacionados = analizar_productos_relacionados_stock(ventas_df, stock_df, catalog_df)
    
    # PRIORIDAD 2: Upselling con stock disponible
    oportunidades_upselling = analizar_upselling_stock_actual(ventas_df, stock_df, catalog_df)
    
    # PRIORIDAD 3: Clientes que barrieron stock (muestra capacidad)
    oportunidades_barren = analizar_clientes_barren_stock(ventas_df, stock_df, catalog_df)
    
    # PRIORIDAD 4: Oportunidades de precio
    oportunidades_precio = analizar_oportunidades_precio(ventas_df, precios_df)
    
    # PRIORIDAD 5: Recompra (menor importancia)
    oportunidades_recompra = analizar_stock_nuevo_recompra(ventas_df, stock_df, catalog_df)
    
    # Crear Excel
    print(f"\n💾 Creando archivo Excel: {OUTPUT_EXCEL}")
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        # Concatenar todas las oportunidades (solo las que tienen datos)
        todas_oportunidades_list = []
        if len(oportunidades_relacionados) > 0:
            todas_oportunidades_list.append(oportunidades_relacionados)
        if len(oportunidades_upselling) > 0:
            todas_oportunidades_list.append(oportunidades_upselling)
        if len(oportunidades_barren) > 0:
            todas_oportunidades_list.append(oportunidades_barren)
        if len(oportunidades_precio) > 0:
            todas_oportunidades_list.append(oportunidades_precio)
        if len(oportunidades_recompra) > 0:
            todas_oportunidades_list.append(oportunidades_recompra)
        
        if len(todas_oportunidades_list) > 0:
            todas_oportunidades = pd.concat(todas_oportunidades_list, ignore_index=True)
            
            # Calcular totales únicos
            try:
                if 'SKU Comprado (Barrió)' in todas_oportunidades.columns:
                    total_oportunidades_unicas = len(todas_oportunidades.drop_duplicates(subset=['Email Cliente', 'SKU Comprado (Barrió)']))
                    total_skus_unicos = len(todas_oportunidades['SKU Comprado (Barrió)'].unique())
                elif 'SKU' in todas_oportunidades.columns:
                    total_oportunidades_unicas = len(todas_oportunidades.drop_duplicates(subset=['Email Cliente', 'SKU']))
                    total_skus_unicos = len(todas_oportunidades['SKU'].unique())
                else:
                    total_oportunidades_unicas = len(todas_oportunidades.drop_duplicates(subset=['Email Cliente']))
                    total_skus_unicos = 0
            except:
                total_oportunidades_unicas = len(todas_oportunidades)
                total_skus_unicos = 0
        else:
            todas_oportunidades = pd.DataFrame()
            total_oportunidades_unicas = 0
            total_skus_unicos = 0
        
        resumen_data = {
            'Métrica': [
                '🎯 PRODUCTOS RELACIONADOS (Stock Disponible)',
                '🎯 UPSELLING (Stock Disponible)',
                '📊 Clientes que Barren Stock (Capacidad)',
                '💰 Oportunidades de Precio',
                '🔄 Recompra (Menor Prioridad)',
                '',
                'Total Oportunidades Únicas',
                'Total Clientes Únicos',
                'Total SKUs Únicos',
                '',
                '⚠️ ENFOQUE: VENDER STOCK ACTUAL',
                '',
                'Este análisis identifica oportunidades comerciales tipo "sniper" para VENDER EL STOCK DISPONIBLE.',
                'NO está enfocado en reposición, sino en maximizar ventas del inventario actual.',
                '',
                'Prioridades:',
                '1. Productos relacionados/familia en stock (cliente barrió uno, hay otros relacionados)',
                '2. Upselling con stock disponible (cliente compró fuerte, hay complementarios)',
                '3. Clientes que barrieron stock (muestra capacidad de compra)',
                '4. Oportunidades de precio (compró más barato que actual)',
                '5. Recompra (menor importancia - enfoque en stock actual)'
            ],
            'Cantidad': [
                len(oportunidades_relacionados),
                len(oportunidades_upselling),
                len(oportunidades_barren),
                len(oportunidades_precio),
                len(oportunidades_recompra),
                '',
                total_oportunidades_unicas,
                len(todas_oportunidades['Email Cliente'].unique()) if len(todas_oportunidades) > 0 else 0,
                total_skus_unicos,
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                ''
            ]
        }
        # Asegurar que ambos arrays tengan la misma longitud
        metricas = resumen_data['Métrica']
        cantidades = resumen_data['Cantidad']
        max_len = max(len(metricas), len(cantidades))
        metricas = metricas + [''] * (max_len - len(metricas))
        cantidades = cantidades + [''] * (max_len - len(cantidades))
        resumen_data = {'Métrica': metricas, 'Cantidad': cantidades}
        
        try:
            resumen_df = pd.DataFrame(resumen_data)
            resumen_df.to_excel(writer, sheet_name='00_Resumen Ejecutivo', index=False)
            auto_adjust_column_widths(writer, '00_Resumen Ejecutivo', resumen_df)
        except Exception as e:
            print(f"   ⚠️  Error creando resumen: {e}")
            # Crear resumen simple si falla
            pd.DataFrame({'Métrica': ['Error al generar resumen'], 'Cantidad': [str(e)]}).to_excel(
                writer, sheet_name='00_Resumen Ejecutivo', index=False
            )
        
        # PRIORIDAD 1: Productos Relacionados (Stock Disponible)
        if len(oportunidades_relacionados) > 0:
            oportunidades_relacionados.to_excel(writer, sheet_name='01_Productos Relacionados (Stock)', index=False)
            auto_adjust_column_widths(writer, '01_Productos Relacionados (Stock)', oportunidades_relacionados)
        
        # PRIORIDAD 2: Upselling (Stock Disponible) - Por ahora igual que relacionados
        if len(oportunidades_upselling) > 0:
            oportunidades_upselling.to_excel(writer, sheet_name='02_Upselling (Stock)', index=False)
            auto_adjust_column_widths(writer, '02_Upselling (Stock)', oportunidades_upselling)
        
        # PRIORIDAD 3: Clientes que Barren Stock (Capacidad)
        if len(oportunidades_barren) > 0:
            oportunidades_barren.to_excel(writer, sheet_name='03_Barren Stock (Capacidad)', index=False)
            auto_adjust_column_widths(writer, '03_Barren Stock (Capacidad)', oportunidades_barren)
            
            # Resumen por rubro y marca
            resumen_barren = generar_resumen_por_rubro_marca(oportunidades_barren)
            if len(resumen_barren) > 0:
                resumen_barren.to_excel(writer, sheet_name='04_Resumen Barren Stock (Rubro-Marca)', index=False)
                auto_adjust_column_widths(writer, '04_Resumen Barren Stock (Rubro-Marca)', resumen_barren)
        
        # PRIORIDAD 4: Oportunidades de Precio
        if len(oportunidades_precio) > 0:
            oportunidades_precio.to_excel(writer, sheet_name='05_Oportunidades Precio', index=False)
            auto_adjust_column_widths(writer, '05_Oportunidades Precio', oportunidades_precio)
        
        # PRIORIDAD 5: Recompra (Menor Importancia)
        if len(oportunidades_recompra) > 0:
            oportunidades_recompra.to_excel(writer, sheet_name='06_Recompra (Menor Prioridad)', index=False)
            auto_adjust_column_widths(writer, '06_Recompra (Menor Prioridad)', oportunidades_recompra)
        
        # Hoja Final: Todas las Oportunidades (Vista Sniper) - SIEMPRE crear esta hoja
        if len(todas_oportunidades) > 0:
            # Ordenar por prioridad y facturación histórica
            if 'Prioridad' in todas_oportunidades.columns:
                todas_oportunidades['Prioridad_Num'] = todas_oportunidades['Prioridad'].map({'ALTA': 1, 'MEDIA': 2, 'BAJA': 3}).fillna(2)
                todas_oportunidades = todas_oportunidades.sort_values(['Prioridad_Num', 'Total Facturación Histórica (USD)'], ascending=[True, False])
                todas_oportunidades = todas_oportunidades.drop('Prioridad_Num', axis=1)
            
            todas_oportunidades.to_excel(writer, sheet_name='07_Todas Oportunidades Sniper', index=False)
            auto_adjust_column_widths(writer, '07_Todas Oportunidades Sniper', todas_oportunidades)
    
    print(f"\n✅ Reporte generado: {OUTPUT_EXCEL}")
    print(f"\n📊 Resumen (ENFOQUE: VENDER STOCK ACTUAL):")
    print(f"   🎯 Productos Relacionados (Stock Disponible): {len(oportunidades_relacionados)}")
    print(f"   🎯 Upselling (Stock Disponible): {len(oportunidades_upselling)}")
    print(f"   📊 Clientes que barren stock (capacidad): {len(oportunidades_barren)}")
    print(f"   💰 Oportunidades de precio: {len(oportunidades_precio)}")
    print(f"   🔄 Recompra (menor prioridad): {len(oportunidades_recompra)}")
    print(f"   - Total oportunidades: {len(todas_oportunidades)}")


if __name__ == "__main__":
    if not HAS_OPENPYXL:
        print("❌ Error: openpyxl no está instalado. Ejecuta: pip install openpyxl")
        exit(1)
    
    generate_sniper_report()
